#!/usr/bin/env python3
"""Intra-sheet edits to an .xlsx workbook with openpyxl.

Set cell values and formulas, insert/delete rows and columns, apply styling,
number formats, and column widths. Sheet-level structure (add/rename/remove/
move/copy a sheet, CSV import) lives in sheets.py -- this script does not
duplicate it.

Round-trip caveat: openpyxl is a high-level model, not a lossless XML editor, so
loading and re-saving a workbook can drop parts it does not represent (pivot
tables, slicers, form controls, VBA, some charts). Every edit therefore writes to
a NEW file (-o; never in place), warns when the input holds such parts, and
refuses to write macro-enabled .xlsm/.xltm. openpyxl also does NOT rewrite
formula references when rows/columns are inserted or deleted, so existing
formulas can point at the wrong cells afterward; those commands print a reminder
to verify them.

Usage:
    kowork-python edit_xlsx.py set <in.xlsx> -o <out.xlsx> --sheet S (--cell A1 | --range A1:B3) [--text] VALUE
    kowork-python edit_xlsx.py insert-rows <in.xlsx> -o <out.xlsx> --sheet S --at N [--count K]
    kowork-python edit_xlsx.py delete-rows <in.xlsx> -o <out.xlsx> --sheet S --at N [--count K]
    kowork-python edit_xlsx.py insert-cols <in.xlsx> -o <out.xlsx> --sheet S --at COL [--count K]
    kowork-python edit_xlsx.py delete-cols <in.xlsx> -o <out.xlsx> --sheet S --at COL [--count K]
    kowork-python edit_xlsx.py style <in.xlsx> -o <out.xlsx> --sheet S --range A1:B2 [--bold] [--italic] [--font NAME] [--size N] [--color RRGGBB] [--fill RRGGBB] [--align left|center|right]
    kowork-python edit_xlsx.py number-format <in.xlsx> -o <out.xlsx> --sheet S --range A1:B2 --format "$#,##0"
    kowork-python edit_xlsx.py width <in.xlsx> -o <out.xlsx> --sheet S --col COL --width N
"""

from __future__ import annotations

import argparse
import sys
from copy import copy

from openpyxl.styles import Color, Font, PatternFill

from xlsxutil import (
    coerce_scalar,
    col_to_index,
    index_to_col,
    load,
    parse_range,
    refuse_inplace,
    refuse_macro_output,
    resolve_sheet,
    warn_lossy_parts,
)

FORMULA_NOTE = (
    "note: openpyxl does not adjust formula references when rows or columns are "
    "inserted or deleted, so existing formulas may now point at the wrong cells "
    "-- verify them, then let Excel recompute on open.\n"
)


class EditError(Exception):
    """A user-facing failure; main prints it as ``error: ...`` and exits 1."""


def write_value(cell, raw: str, force_text: bool) -> None:
    """Write VALUE into a cell: forced string, formula (leading '='), or scalar."""
    if not force_text and raw.startswith("="):
        cell.value = raw  # a formula
        return
    cell.value = coerce_scalar(raw, force_text=force_text)
    # openpyxl treats a leading '=' string as a formula; pin to string when the
    # user forced text so a literal "=..." is kept as text.
    if force_text and isinstance(cell.value, str) and cell.value.startswith("="):
        cell.data_type = "s"


def resolve_col(spec) -> int:
    """A column given as a letter (``A``) or a 1-based index (``1``) -> its index."""
    s = str(spec).strip()
    if s.isdigit():
        idx = int(s)
        if idx < 1:
            raise EditError(f"column index must be >= 1, got {spec!r}")
        return idx
    return col_to_index(s)


def norm_color(value: str) -> str:
    """Normalise a colour to openpyxl's 8-hex AARRGGBB (alpha defaults to FF)."""
    v = value.strip().lstrip("#").upper()
    if len(v) == 6:
        v = "FF" + v
    if len(v) != 8:
        raise EditError(f"colour must be 6 (RRGGBB) or 8 (AARRGGBB) hex digits, got {value!r}")
    try:
        int(v, 16)
    except ValueError:
        raise EditError(f"invalid hex colour {value!r}")
    return v


def iter_range(ws, spec):
    """Yield every cell in a range spec (validated, bounded)."""
    min_col, min_row, max_col, max_row = parse_range(spec)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            yield cell


def apply_style(cell, args: argparse.Namespace) -> None:
    """Apply only the style attributes given, preserving the cell's others."""
    if args.bold or args.italic or args.font or args.size is not None or args.color:
        font = copy(cell.font)
        if args.bold:
            font.bold = True
        if args.italic:
            font.italic = True
        if args.font:
            font.name = args.font
        if args.size is not None:
            font.size = args.size
        if args.color:
            font.color = Color(rgb=norm_color(args.color))
        cell.font = font
    if args.fill:
        cell.fill = PatternFill("solid", fgColor=norm_color(args.fill))
    if args.align:
        alignment = copy(cell.alignment)
        alignment.horizontal = args.align
        cell.alignment = alignment


def cmd_set(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    force_text = args.text
    is_formula = (not force_text) and args.value.startswith("=")
    if is_formula and args.range:
        raise EditError(
            "a formula can only be set on a single --cell: openpyxl writes the "
            "formula text verbatim to every cell of a --range without adjusting "
            "relative references. Use --cell for a formula; --range fills constants."
        )

    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)

    if args.cell:
        min_col, min_row, max_col, max_row = parse_range(args.cell)
        if (min_col, min_row) != (max_col, max_row):
            raise EditError("--cell must be a single cell like A1; use --range for a block")
        write_value(ws.cell(row=min_row, column=min_col), args.value, force_text)
        target = args.cell
    else:
        for cell in iter_range(ws, args.range):
            write_value(cell, args.value, force_text)
        target = args.range

    warn_lossy_parts(args.input)
    wb.save(args.out)
    kind = "text" if force_text else ("formula" if is_formula else "value")
    print(f"set {kind} on {ws.title}!{target} -> {args.out}")
    return 0


def cmd_insert_rows(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    if args.at < 1:
        raise EditError("--at must be a 1-based row index >= 1")
    if args.count < 1:
        raise EditError("--count must be >= 1")
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    ws.insert_rows(args.at, amount=args.count)
    sys.stderr.write(FORMULA_NOTE)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"inserted {args.count} row(s) at row {args.at} on {ws.title} -> {args.out}")
    return 0


def cmd_delete_rows(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    if args.at < 1:
        raise EditError("--at must be a 1-based row index >= 1")
    if args.count < 1:
        raise EditError("--count must be >= 1")
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    ws.delete_rows(args.at, amount=args.count)
    sys.stderr.write(FORMULA_NOTE)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"deleted {args.count} row(s) from row {args.at} on {ws.title} -> {args.out}")
    return 0


def cmd_insert_cols(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    if args.count < 1:
        raise EditError("--count must be >= 1")
    idx = resolve_col(args.at)
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    ws.insert_cols(idx, amount=args.count)
    sys.stderr.write(FORMULA_NOTE)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"inserted {args.count} column(s) at column {index_to_col(idx)} on {ws.title} -> {args.out}")
    return 0


def cmd_delete_cols(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    if args.count < 1:
        raise EditError("--count must be >= 1")
    idx = resolve_col(args.at)
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    ws.delete_cols(idx, amount=args.count)
    sys.stderr.write(FORMULA_NOTE)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"deleted {args.count} column(s) from column {index_to_col(idx)} on {ws.title} -> {args.out}")
    return 0


def cmd_style(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    if not any([args.bold, args.italic, args.font, args.size is not None, args.color, args.fill, args.align]):
        raise EditError("provide at least one of --bold/--italic/--font/--size/--color/--fill/--align")
    if args.size is not None and args.size <= 0:
        raise EditError("--size must be > 0")
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    for cell in iter_range(ws, args.range):
        apply_style(cell, args)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"styled {ws.title}!{args.range} -> {args.out}")
    return 0


def cmd_number_format(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    for cell in iter_range(ws, args.range):
        cell.number_format = args.format
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"set number format {args.format!r} on {ws.title}!{args.range} -> {args.out}")
    return 0


def cmd_width(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    if args.width <= 0:
        raise EditError("--width must be > 0")
    letter = index_to_col(resolve_col(args.col))
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    ws.column_dimensions[letter].width = args.width
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"set column {letter} width to {args.width} on {ws.title} -> {args.out}")
    return 0


def _add_io(p: argparse.ArgumentParser) -> None:
    p.add_argument("input", help="path to the .xlsx/.xlsm file")
    p.add_argument("-o", "--out", required=True, help="path to write the result")
    p.add_argument("--sheet", required=True, help="target sheet (name or 1-based index)")


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Intra-sheet edits to an .xlsx (openpyxl).")
    sub = ap.add_subparsers(dest="command", required=True)

    p = sub.add_parser("set", help="set a cell/range value or a single-cell formula")
    _add_io(p)
    where = p.add_mutually_exclusive_group(required=True)
    where.add_argument("--cell", help="a single cell, e.g. A1")
    where.add_argument("--range", help="a block to fill with the constant, e.g. A1:B3")
    p.add_argument("--text", action="store_true", help="force the value to a string (no numeric coercion; keeps a literal '=')")
    p.add_argument("value", help="value to set; a leading '=' is a formula (single cell only) unless --text")
    p.set_defaults(func=cmd_set)

    p = sub.add_parser("insert-rows", help="insert blank rows")
    _add_io(p)
    p.add_argument("--at", type=int, required=True, help="1-based row index to insert before")
    p.add_argument("--count", type=int, default=1, help="how many rows (default: 1)")
    p.set_defaults(func=cmd_insert_rows)

    p = sub.add_parser("delete-rows", help="delete rows")
    _add_io(p)
    p.add_argument("--at", type=int, required=True, help="1-based first row to delete")
    p.add_argument("--count", type=int, default=1, help="how many rows (default: 1)")
    p.set_defaults(func=cmd_delete_rows)

    p = sub.add_parser("insert-cols", help="insert blank columns")
    _add_io(p)
    p.add_argument("--at", required=True, help="column to insert before (letter or 1-based index)")
    p.add_argument("--count", type=int, default=1, help="how many columns (default: 1)")
    p.set_defaults(func=cmd_insert_cols)

    p = sub.add_parser("delete-cols", help="delete columns")
    _add_io(p)
    p.add_argument("--at", required=True, help="first column to delete (letter or 1-based index)")
    p.add_argument("--count", type=int, default=1, help="how many columns (default: 1)")
    p.set_defaults(func=cmd_delete_cols)

    p = sub.add_parser("style", help="apply font/fill/alignment over a range")
    _add_io(p)
    p.add_argument("--range", required=True, help="range to style, e.g. A1:B2")
    p.add_argument("--bold", action="store_true", help="make text bold")
    p.add_argument("--italic", action="store_true", help="make text italic")
    p.add_argument("--font", help="font family name, e.g. Arial")
    p.add_argument("--size", type=float, help="font size in points")
    p.add_argument("--color", help="font colour, RRGGBB or AARRGGBB hex")
    p.add_argument("--fill", help="solid fill colour, RRGGBB or AARRGGBB hex")
    p.add_argument("--align", choices=["left", "center", "right"], help="horizontal alignment")
    p.set_defaults(func=cmd_style)

    p = sub.add_parser("number-format", help="set the number format over a range")
    _add_io(p)
    p.add_argument("--range", required=True, help="range to format, e.g. A1:B2")
    p.add_argument("--format", required=True, help="number format code, e.g. '$#,##0' or '0.0%%'")
    p.set_defaults(func=cmd_number_format)

    p = sub.add_parser("width", help="set a column's width")
    _add_io(p)
    p.add_argument("--col", required=True, help="column to size (letter or 1-based index)")
    p.add_argument("--width", type=float, required=True, help="width in Excel character units")
    p.set_defaults(func=cmd_width)

    return ap


def main(argv: list[str]) -> int:
    args = build_parser().parse_args(argv)
    try:
        return args.func(args)
    except Exception as exc:
        sys.stderr.write(f"error: {exc}\n")
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
