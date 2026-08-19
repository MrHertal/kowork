#!/usr/bin/env python3
"""Authoring template for creating an Excel workbook with openpyxl.

Copy this into a temp directory (never the user's folder), edit the
``build_workbook()`` function to build the requested content, then run it to
write the .xlsx to the path the user wants:

    kowork-python create_xlsx.py out.xlsx

It runs from a temp directory (never the user's folder); the copy is kept there
after a successful write so you can edit it and re-run to revise the workbook in
the same session (the OS reclaims temp later).

It demonstrates every "create" building block, each editable in one obvious
place: a bold, filled header row; data rows; number formats (currency and
percent); live formulas (a per-row ``=B*C`` and a ``=SUM(...)`` total); frozen
panes; column widths; a second sheet (with a cross-sheet formula); a bar chart
built from the data; and an embedded image. The default font is Calibri 11
(Excel's own default); change DEFAULT_FONT_NAME / DEFAULT_FONT_SIZE to switch.
The workbook is saved with the current Office theme, so charts use the same
colors as a new Excel file.

Prefer real Excel formulas over Python-computed constants so the workbook stays
live and recalculates in Excel; only compute a value in Python and write the
static number when the user explicitly needs a fixed result. openpyxl writes
formulas but never evaluates them, so a formula cell has no cached value until
Excel opens and saves the file -- reading it back with data_only=True returns
None for those cells, which is not the same as empty.

Image note: the demo PNG is generated in memory so the template needs no
external asset. For a real picture, point ``demo_image()`` at a file with
``XLImage("photo.png")``; to omit the image, delete the ``ws.add_image(...)``
line in build_workbook().

Macro-enabled output (.xlsm / .xltm) is refused: openpyxl can silently drop
macros and other parts it does not model, so authoring those is out of scope --
save a plain .xlsx instead.

Usage:
    kowork-python create_xlsx.py <out.xlsx>
"""

from __future__ import annotations

import argparse
import io
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage

DEFAULT_FONT_NAME = "Calibri"
DEFAULT_FONT_SIZE = 11
HEADER_FILL = "FF34507A"
HEADER_FONT_COLOR = "FFFFFFFF"
CURRENCY_FORMAT = "$#,##0"
PERCENT_FORMAT = "0.0%"


def demo_image() -> XLImage:
    """A tiny solid-colour PNG built in memory, so the template needs no assets.

    Swap for ``XLImage("photo.png")`` to embed a real file.
    """
    buf = io.BytesIO()
    PILImage.new("RGB", (96, 96), (52, 80, 122)).save(buf, format="PNG")
    buf.seek(0)
    return XLImage(buf)


def build_workbook(wb: Workbook) -> None:
    """The workbook content. Edit this function to change what the file holds."""
    # Unstyled cells inherit openpyxl's built-in default font, which is already
    # Calibri 11 (matching DEFAULT_FONT_*). openpyxl has no public default-font
    # setter, so to restyle the sheet you set Font(...) on the cells you create
    # -- as the header and total below do with the DEFAULT_FONT_* constants.

    ws = wb.active
    ws.title = "Sales"

    header = ["Item", "Quantity", "Unit Price", "Amount", "Margin"]
    data_rows = [
        ("Widget", 30, 4, 0.25),
        ("Gadget", 12, 9, 0.40),
        ("Gizmo", 7, 25, 0.55),
    ]

    # Header row: bold white text on a solid fill, centred.
    ws.append(header)
    header_font = Font(
        name=DEFAULT_FONT_NAME,
        size=DEFAULT_FONT_SIZE,
        bold=True,
        color=HEADER_FONT_COLOR,
    )
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows. Amount (column D) is a live per-row formula =B*C rather than a
    # Python product, so Excel recalculates it; Margin (column E) is a percent.
    first_data_row = 2
    for offset, (item, qty, price, margin) in enumerate(data_rows):
        r = first_data_row + offset
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=3, value=price).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=5, value=margin).number_format = PERCENT_FORMAT

    # Total row: a live SUM over the Amount column (not a precomputed number).
    last_data_row = first_data_row + len(data_rows) - 1
    total_row = last_data_row + 1
    bold = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    label_cell = ws.cell(row=total_row, column=1, value="Total")
    label_cell.font = bold
    total_cell = ws.cell(
        row=total_row,
        column=4,
        value=f"=SUM(D{first_data_row}:D{last_data_row})",
    )
    total_cell.font = bold
    total_cell.number_format = CURRENCY_FORMAT

    # Freeze the header row so it stays visible while scrolling.
    ws.freeze_panes = "A2"

    # Column widths, in Excel's character-width units.
    for col_letter, width in {"A": 16, "B": 10, "C": 12, "D": 12, "E": 10}.items():
        ws.column_dimensions[col_letter].width = width

    # Bar chart of Amount by Item, built from the cells above. The Amount
    # reference includes the header row so titles_from_data names the series.
    chart = BarChart()
    chart.type = "col"
    chart.title = "Amount by item"
    chart.x_axis.title = "Item"
    chart.y_axis.title = "Amount"
    amounts = Reference(ws, min_col=4, min_row=1, max_row=last_data_row)
    items = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(amounts, titles_from_data=True)
    chart.set_categories(items)
    ws.add_chart(chart, "G2")

    # Embedded image (REMOVABLE -- delete this line to omit it).
    ws.add_image(demo_image(), "G18")

    # A second sheet, carrying a cross-sheet formula back to the total above.
    summary = wb.create_sheet("Summary")
    summary.cell(row=1, column=1, value="Total amount").font = bold
    summary.cell(row=1, column=2, value=f"='Sales'!D{total_row}").number_format = CURRENCY_FORMAT
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 12


def modern_office_theme() -> str:
    """openpyxl's built-in theme with the current Office palette and heading font.

    The bundled ``theme_xml`` predates the current Office colors, so charts
    (which take their series colors from the theme) would render in the old
    palette. Swap in the current values; every replaced string occurs exactly
    once in ``theme_xml``.
    """
    from openpyxl.writer.theme import theme_xml

    theme = theme_xml
    for old, new in {
        'typeface="Cambria"': 'typeface="Calibri Light"',  # majorFont (minorFont is already Calibri)
        "4F81BD": "4472C4",  # accent1
        "C0504D": "ED7D31",  # accent2
        "9BBB59": "A5A5A5",  # accent3
        "8064A2": "FFC000",  # accent4
        "4BACC6": "5B9BD5",  # accent5
        "F79646": "70AD47",  # accent6
        "1F497D": "44546A",  # dk2
        "EEECE1": "E7E6E6",  # lt2
        "0000FF": "0563C1",  # hyperlink
        "800080": "954F72",  # followed hyperlink
    }.items():
        assert theme.count(old) == 1, f"expected exactly one {old} in openpyxl's theme_xml"
        theme = theme.replace(old, new)
    return theme


def build_xlsx(out_path: str) -> None:
    wb = Workbook()
    build_workbook(wb)
    wb.loaded_theme = modern_office_theme()
    wb.save(out_path)


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description="Create an .xlsx workbook from an editable openpyxl template.")
    ap.add_argument("output", nargs="?", default="output.xlsx", help="path to write the .xlsx")
    args = ap.parse_args(argv)

    ext = os.path.splitext(args.output)[1].lower()
    if ext in (".xlsm", ".xltm"):
        sys.stderr.write(
            f"error: refusing to write a macro-enabled workbook ({ext}); openpyxl "
            "can silently drop macros and other parts it does not model. Write a "
            ".xlsx instead.\n"
        )
        return 1

    try:
        build_xlsx(args.output)
    except Exception as exc:
        sys.stderr.write(f"error: {exc}\n")
        return 1

    size = os.path.getsize(args.output)
    print(f"wrote {args.output} {size} bytes")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
