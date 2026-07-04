#!/usr/bin/env python3
"""Workbook and sheet structure + metadata for .xlsx files, with openpyxl.

One tool with a subcommand per operation. ``info`` is a read-only structural
summary; the rest add, rename, remove, move, copy, or import-from-CSV a sheet and
write the result to a new file. This script owns sheet-level structure only --
editing cell values, rows/columns, and styles lives in edit_xlsx.py.

Rules (consistent across the xlsx skill):
  * Mutating subcommands always write to -o; they never edit in place, and they
    refuse to overwrite the input (openpyxl round-trips can be lossy).
  * Mutating subcommands refuse to write .xlsm/.xltm: openpyxl can silently drop
    macros and other parts it does not model. Save a plain .xlsx instead.
  * Before saving a result derived from an existing workbook, the input is
    scanned for parts openpyxl may drop (pivots/slicers/form-controls/VBA) and a
    non-fatal warning is printed.

Usage:
    kowork-python sheets.py info <in.xlsx>
    kowork-python sheets.py add <in.xlsx> -o <out.xlsx> --name NAME [--index N]
    kowork-python sheets.py rename <in.xlsx> -o <out.xlsx> --sheet OLD --to NEW
    kowork-python sheets.py remove <in.xlsx> -o <out.xlsx> --sheet NAME
    kowork-python sheets.py move <in.xlsx> -o <out.xlsx> --sheet NAME --to-index N
    kowork-python sheets.py copy <in.xlsx> -o <out.xlsx> --sheet NAME [--to NEWNAME]
    kowork-python sheets.py from-csv <in.csv> -o <out.xlsx> [--into existing.xlsx] [--sheet NAME] [--delimiter ,] [--text-columns A,C]
"""

from __future__ import annotations

import argparse
import csv
import os
import sys

from openpyxl import Workbook

from xlsxutil import (
    coerce_scalar,
    col_to_index,
    format_table,
    load,
    refuse_inplace,
    refuse_macro_output,
    resolve_sheet,
    warn_lossy_parts,
)

MAX_SHEET_TITLE = 31  # Excel's hard limit on a sheet name's length.


class SheetsError(Exception):
    """A user-facing failure; main prints it as ``error: ...`` and exits 1."""


def keep_active(wb, active_title: str) -> None:
    """Re-point ``wb.active`` to ``active_title`` after a structural change.

    openpyxl tracks the active sheet by position, so moving or removing a sheet
    can silently change which one is active. Re-point to the recorded title's new
    index; if that title is gone (it was the removed sheet), fall back to the
    first remaining sheet.
    """
    names = wb.sheetnames
    target = active_title if active_title in names else names[0]
    wb.active = names.index(target)


def parse_text_columns(spec):
    """Parse a ``--text-columns`` spec ('A,C' or '1,3') into a set of 1-based indices."""
    cols = set()
    for token in (spec or "").split(","):
        token = token.strip()
        if not token:
            continue
        if token.isdigit():
            idx = int(token)
            if idx < 1:
                raise SheetsError(f"--text-columns index must be >= 1, got {token!r}")
            cols.add(idx)
        else:
            cols.add(col_to_index(token))
    return cols


def _defined_names(wb):
    """(name, reference) pairs for the workbook's defined names, API-version safe."""
    dn = wb.defined_names
    try:
        items = list(dn.items())  # dict-like in openpyxl >= 3.1
    except AttributeError:
        items = [(getattr(d, "name", str(d)), d) for d in dn]
    return [(name, getattr(obj, "value", obj)) for name, obj in items]


def cmd_info(args: argparse.Namespace) -> int:
    # A normal (non-streaming) load, so merged cells, charts, and images are
    # populated -- read_only worksheets do not expose them. A very large
    # workbook is therefore read fully into memory here.
    wb = load(args.input)
    print(f"workbook: {args.input}")
    print(f"sheets: {len(wb.sheetnames)} | active: {wb.active.title!r}")

    rows = [["Sheet", "Dimensions", "Rows", "Cols", "Merged", "Freeze", "Charts", "Images"]]
    merged_notes = []
    for ws in wb.worksheets:
        merged = list(ws.merged_cells.ranges)
        rows.append([
            ws.title,
            ws.dimensions,
            ws.max_row,
            ws.max_column,
            len(merged),
            ws.freeze_panes or "(none)",
            len(ws._charts),
            len(ws._images),
        ])
        if merged:
            shown = ", ".join(sorted(str(r) for r in merged)[:5])
            extra = f" (+{len(merged) - 5} more)" if len(merged) > 5 else ""
            merged_notes.append(f"  {ws.title!r}: {shown}{extra}")

    print()
    print(format_table(rows))

    if merged_notes:
        print()
        print("merged ranges:")
        for line in merged_notes:
            print(line)

    names = _defined_names(wb)
    print()
    if names:
        print("defined names:")
        for name, ref in names:
            print(f"  {name} = {ref}")
    else:
        print("defined names: (none)")
    return 0


def cmd_add(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    wb = load(args.input)
    if args.name in wb.sheetnames:
        raise SheetsError(f"a sheet named {args.name!r} already exists")
    if args.index is not None and not (0 <= args.index <= len(wb.sheetnames)):
        raise SheetsError(f"--index {args.index} out of range 0-{len(wb.sheetnames)}")
    active_title = wb.active.title
    wb.create_sheet(title=args.name, index=args.index)
    # openpyxl tracks the active sheet by position, so inserting at or before it
    # would silently make the new sheet active; keep the original active sheet.
    keep_active(wb, active_title)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    where = f" at index {args.index}" if args.index is not None else ""
    print(f"added sheet {args.name!r}{where} -> {args.out}")
    return 0


def cmd_rename(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    if args.to in wb.sheetnames and args.to != ws.title:
        raise SheetsError(f"a sheet named {args.to!r} already exists")
    old = ws.title
    ws.title = args.to
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"renamed {old!r} to {args.to!r} -> {args.out}")
    return 0


def cmd_remove(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    if len(wb.sheetnames) <= 1:
        raise SheetsError("cannot remove the only sheet in the workbook")
    active_title = wb.active.title
    title = ws.title
    wb.remove(ws)
    keep_active(wb, active_title)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"removed sheet {title!r} -> {args.out}")
    return 0


def cmd_move(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    count = len(wb.sheetnames)
    if not (0 <= args.to_index < count):
        raise SheetsError(f"--to-index {args.to_index} out of range 0-{count - 1}")
    active_title = wb.active.title
    # openpyxl moves by a relative offset, so translate the absolute target.
    offset = args.to_index - wb.sheetnames.index(ws.title)
    wb.move_sheet(ws.title, offset=offset)
    keep_active(wb, active_title)
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"moved sheet {ws.title!r} to index {args.to_index} -> {args.out}")
    return 0


def cmd_copy(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    refuse_inplace(args.out, args.input)
    wb = load(args.input)
    ws = resolve_sheet(wb, args.sheet)
    new_title = args.to or f"{ws.title} Copy"
    if new_title in wb.sheetnames:
        raise SheetsError(f"a sheet named {new_title!r} already exists")
    # copy_worksheet duplicates cell values and styles but not charts, images,
    # data validations, or conditional formatting (an openpyxl limitation).
    new_ws = wb.copy_worksheet(ws)
    new_ws.title = new_title
    dropped = []
    if ws._charts:
        dropped.append(f"{len(ws._charts)} chart(s)")
    if ws._images:
        dropped.append(f"{len(ws._images)} image(s)")
    if ws.data_validations.dataValidation:
        dropped.append(f"{len(ws.data_validations.dataValidation)} data validation(s)")
    cf_count = len(list(ws.conditional_formatting))
    if cf_count:
        dropped.append(f"{cf_count} conditional-formatting rule(s)")
    if dropped:
        sys.stderr.write(
            f"warning: the copy of {ws.title!r} will not include its "
            f"{', '.join(dropped)}; openpyxl's copy_worksheet duplicates only "
            "cell values and styles.\n"
        )
    warn_lossy_parts(args.input)
    wb.save(args.out)
    print(f"copied sheet {ws.title!r} to {new_title!r} -> {args.out}")
    return 0


def cmd_from_csv(args: argparse.Namespace) -> int:
    refuse_macro_output(args.out)
    if not os.path.isfile(args.input):
        raise SheetsError(f"no such file: {args.input}")
    refuse_inplace(args.out, args.input)
    if len(args.delimiter) != 1:
        raise SheetsError("--delimiter must be a single character")

    text_cols = parse_text_columns(args.text_columns)
    with open(args.input, newline="", encoding="utf-8") as fh:
        rows = [
            [coerce_scalar(v, force_text=j in text_cols) for j, v in enumerate(row, start=1)]
            for row in csv.reader(fh, delimiter=args.delimiter)
        ]

    sheet_name = (args.sheet or os.path.splitext(os.path.basename(args.input))[0])[:MAX_SHEET_TITLE]

    if args.into:
        refuse_inplace(args.out, args.into)
        wb = load(args.into)
        if sheet_name in wb.sheetnames:
            raise SheetsError(f"a sheet named {sheet_name!r} already exists in {args.into}")
        ws = wb.create_sheet(title=sheet_name)
        warn_lossy_parts(args.into)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    for row in rows:
        ws.append(row)
    # CSV cells are data, not formulas: openpyxl turns any string beginning with
    # '=' into a formula, so pin those back to literal text. A real CSV value
    # like "=1+1" stays the text "=1+1" (and --text-columns is honoured for it).
    for r in ws.iter_rows():
        for c in r:
            if isinstance(c.value, str) and c.value.startswith("="):
                c.data_type = "s"
    wb.save(args.out)
    print(f"wrote {len(rows)} row(s) to sheet {sheet_name!r} -> {args.out}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Workbook and sheet structure + metadata (openpyxl).")
    sub = ap.add_subparsers(dest="command", required=True)

    p = sub.add_parser("info", help="print workbook + per-sheet structure")
    p.add_argument("input", help="path to the .xlsx/.xlsm file")
    p.set_defaults(func=cmd_info)

    p = sub.add_parser("add", help="add a new empty sheet")
    p.add_argument("input", help="path to the .xlsx/.xlsm file")
    p.add_argument("-o", "--out", required=True, help="path to write the result")
    p.add_argument("--name", required=True, help="name of the new sheet")
    p.add_argument("--index", type=int, help="0-based position (default: append at the end)")
    p.set_defaults(func=cmd_add)

    p = sub.add_parser("rename", help="rename a sheet")
    p.add_argument("input", help="path to the .xlsx/.xlsm file")
    p.add_argument("-o", "--out", required=True, help="path to write the result")
    p.add_argument("--sheet", required=True, help="sheet to rename (name or 1-based index)")
    p.add_argument("--to", required=True, help="new sheet name")
    p.set_defaults(func=cmd_rename)

    p = sub.add_parser("remove", help="remove a sheet (not the last remaining one)")
    p.add_argument("input", help="path to the .xlsx/.xlsm file")
    p.add_argument("-o", "--out", required=True, help="path to write the result")
    p.add_argument("--sheet", required=True, help="sheet to remove (name or 1-based index)")
    p.set_defaults(func=cmd_remove)

    p = sub.add_parser("move", help="reposition a sheet")
    p.add_argument("input", help="path to the .xlsx/.xlsm file")
    p.add_argument("-o", "--out", required=True, help="path to write the result")
    p.add_argument("--sheet", required=True, help="sheet to move (name or 1-based index)")
    p.add_argument("--to-index", dest="to_index", type=int, required=True, help="0-based destination position")
    p.set_defaults(func=cmd_move)

    p = sub.add_parser("copy", help="duplicate a sheet within the workbook")
    p.add_argument("input", help="path to the .xlsx/.xlsm file")
    p.add_argument("-o", "--out", required=True, help="path to write the result")
    p.add_argument("--sheet", required=True, help="sheet to copy (name or 1-based index)")
    p.add_argument("--to", help="name for the copy (default: '<name> Copy')")
    p.set_defaults(func=cmd_copy)

    p = sub.add_parser("from-csv", help="import a CSV file as a sheet")
    p.add_argument("input", help="path to the .csv file")
    p.add_argument("-o", "--out", required=True, help="path to write the .xlsx")
    p.add_argument("--into", help="append as a new sheet to this existing .xlsx (default: new workbook)")
    p.add_argument("--sheet", help="name for the new sheet (default: the CSV's basename)")
    p.add_argument("--delimiter", default=",", help="CSV field delimiter (default: ',')")
    p.add_argument("--text-columns", dest="text_columns", help="columns to keep as raw text, no numeric coercion, e.g. 'A,C' or '1,3'")
    p.set_defaults(func=cmd_from_csv)

    return ap


def main(argv: list[str]) -> int:
    args = build_parser().parse_args(argv)
    try:
        return args.func(args)
    except Exception as exc:
        sys.stderr.write(f"error: {exc}\n")
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
