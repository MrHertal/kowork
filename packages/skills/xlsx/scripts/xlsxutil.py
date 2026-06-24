"""Shared helpers for the xlsx skill.

Small, dependency-free utilities the xlsx scripts have in common, kept here so
the individual scripts stay short. Workbook loading, sheet/range/column parsing,
the Markdown-table formatter, and the round-trip-fidelity warning all live here
as the single source of truth, imported as ``from xlsxutil import ...`` -- the
script's own directory is on ``sys.path`` when run as ``kowork-python
scripts/<name>.py``.

``create_xlsx.py`` deliberately does NOT import this module: it is copied alone
into a temp directory before being run, so it must stay self-contained.
"""

from __future__ import annotations

import math
import os
import sys
import zipfile

import openpyxl
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)
from openpyxl.utils.exceptions import InvalidFileException

# Extensions openpyxl can read. .xls / .ods are different formats it cannot open.
READABLE_EXTS = (".xlsx", ".xlsm", ".xltx")

# Macro-enabled outputs the mutating scripts refuse to write: openpyxl can
# silently drop macros and other parts it does not model.
MACRO_EXTS = (".xlsm", ".xltm")


class XlsxError(Exception):
    """A user-facing failure; callers print it as ``error: ...`` and exit 1."""


def load(path, *, read_only=False, data_only=False):
    """Open a workbook, turning the common failures into a clean ``XlsxError``.

    ``read_only`` streams the file (bounded memory for large books); ``data_only``
    returns cached values instead of formulas. Reading an ``.xlsm`` is fine (its
    macros are simply ignored on read) -- this helper never writes, and the
    mutating scripts refuse to *save* ``.xlsm``.
    """
    if not os.path.isfile(path):
        raise XlsxError(f"no such file: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext not in READABLE_EXTS:
        raise XlsxError(
            f"unsupported file type {ext or '(none)'}: expected one of "
            f"{', '.join(READABLE_EXTS)} (openpyxl cannot read .xls or .ods)"
        )
    try:
        return openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)
    except InvalidFileException as exc:
        raise XlsxError(f"not a readable .xlsx workbook: {exc}")
    except zipfile.BadZipFile:
        raise XlsxError(f"not a valid .xlsx file (not a zip): {path}")
    except Exception as exc:
        raise XlsxError(f"cannot open {path}: {exc}")


def refuse_macro_output(out_path):
    """Reject a macro-enabled output path; the mutating scripts never write one."""
    ext = os.path.splitext(out_path)[1].lower()
    if ext in MACRO_EXTS:
        raise XlsxError(
            f"refusing to write a macro-enabled workbook ({ext}); openpyxl can "
            "silently drop macros and other parts it does not model. Write a "
            ".xlsx instead."
        )


def refuse_inplace(out_path, source_path):
    """Block writing the result back over a file we are reading from."""
    try:
        same = os.path.realpath(out_path) == os.path.realpath(source_path)
    except OSError:
        same = os.path.abspath(out_path) == os.path.abspath(source_path)
    if same:
        raise XlsxError(
            f"refusing to overwrite {source_path} in place; openpyxl round-trips "
            "can be lossy, so write to a different -o path"
        )


def coerce_scalar(raw, *, force_text=False):
    """Light typing for a single value: stays a string unless it parses cleanly
    as an int or a finite float; a blank becomes ``None`` (an empty cell).

    ``force_text`` keeps the raw string verbatim (no numeric coercion), so a
    numeric-looking value such as "007" survives -- used by ``set --text`` and
    ``from-csv --text-columns``. Without it, "007" becomes the number 7, matching
    Excel's own import.
    """
    if raw is None:
        return None
    s = raw.strip()
    if not s:
        return None
    if force_text:
        return raw
    try:
        return int(s)
    except ValueError:
        pass
    try:
        f = float(s)
    except ValueError:
        return raw
    return f if math.isfinite(f) else raw


def resolve_sheet(wb, name=None):
    """Return the worksheet for a ``--sheet`` value (a name or a 1-based index).

    Defaults to the active sheet. An exact sheet-name match wins; otherwise a
    bare integer is treated as a 1-based position. A bad value raises
    ``XlsxError`` listing the available sheet names.
    """
    if name is None:
        return wb.active
    key = str(name)
    if key in wb.sheetnames:
        return wb[key]
    if key.isdigit():
        idx = int(key)
        if 1 <= idx <= len(wb.sheetnames):
            return wb.worksheets[idx - 1]
        raise XlsxError(f"sheet index {idx} out of range 1-{len(wb.sheetnames)}")
    available = ", ".join(repr(n) for n in wb.sheetnames)
    raise XlsxError(f"no sheet named {key!r}; available sheets: {available}")


def parse_range(spec):
    """Parse ``A1:D10`` into ``(min_col, min_row, max_col, max_row)`` (1-based).

    Wraps openpyxl's ``range_boundaries`` with validation: a single cell (``A1``)
    is a 1x1 range; an open-ended range (``A:D`` or ``1:10``) is rejected because
    the read paths need explicit row and column bounds.
    """
    try:
        min_col, min_row, max_col, max_row = range_boundaries(str(spec))
    except (ValueError, TypeError):
        raise XlsxError(f"invalid range {spec!r}; use a form like A1 or A1:D10")
    if None in (min_col, min_row, max_col, max_row):
        raise XlsxError(f"range {spec!r} must be bounded (e.g. A1:D10), not open-ended")
    return min_col, min_row, max_col, max_row


def col_to_index(letter):
    """Column letter (``A``) to its 1-based index; ``XlsxError`` on bad input."""
    try:
        return column_index_from_string(str(letter).strip().upper())
    except (ValueError, TypeError):
        raise XlsxError(f"invalid column letter {letter!r}")


def index_to_col(idx):
    """1-based column index to its letter (``A``); ``XlsxError`` on bad input."""
    try:
        return get_column_letter(int(idx))
    except (ValueError, TypeError):
        raise XlsxError(f"invalid column index {idx!r}")


def cell(value):
    """One value as a single-line, pipe-safe string for a Markdown table cell."""
    if value is None:
        return ""
    return " ".join(str(value).split()).replace("|", "\\|")


def format_table(rows):
    """Render ``rows`` (an iterable of iterables) as a GitHub-style pipe table.

    The first row is the header, followed by a separator line. Ragged rows are
    padded to the widest row. Returns ``""`` when there are no rows.
    """
    cleaned = [[cell(c) for c in row] for row in rows]
    width = max((len(r) for r in cleaned), default=0)
    if width == 0:
        return ""
    for r in cleaned:
        r.extend([""] * (width - len(r)))
    header, *body = cleaned
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def warn_lossy_parts(path, *, editing_charts=False):
    """Warn once (to stderr) if the input workbook holds parts openpyxl may drop.

    openpyxl is a high-level model, not a lossless XML editor, so on load+save it
    can silently discard parts it does not represent. Before a mutating script
    saves a result derived from ``path``, peek at the package's zip namelist and
    flag any such parts. Charts are only flagged when ``editing_charts`` is set,
    since a chart-touching edit is the case where they are at risk. Returns the
    list of part labels found -- empty if none, or if the file cannot be opened
    as a zip (``load()`` surfaces real open errors).
    """
    try:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
    except (OSError, zipfile.BadZipFile):
        return []
    checks = [
        ("pivot tables", lambda n: n.startswith("xl/pivotTables/")),
        ("slicers", lambda n: n.startswith("xl/slicers") or n.startswith("xl/slicerCaches/")),
        ("form controls", lambda n: n.startswith("xl/ctrlProps/")),
        ("VBA macros", lambda n: n.endswith("vbaProject.bin")),
    ]
    if editing_charts:
        checks.append(("charts", lambda n: n.startswith("xl/charts/")))
    found = [label for label, pred in checks if any(pred(n) for n in names)]
    if found:
        parts = ", ".join(found)
        sys.stderr.write(
            f"warning: {path} contains {parts} that openpyxl may drop when "
            "saving the result; verify by reopening, or edit a copy in Excel.\n"
        )
    return found
